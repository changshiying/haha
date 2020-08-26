# import openpyxl
# # wb=openpyxl.load_workbook('test_case_api.xlsx')
# # sheet=wb['register']
# # sheet.cell(row=2,column=8).value='pass'
# # wb.save('test_case_api.xlsx')


import openpyxl
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)
    sheet=wb['sheetname']
    sheet.cell(row=row,column=column).value=filename
    wb.sava('test_case_api.xlsx')
